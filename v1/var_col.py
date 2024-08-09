### Extracts dataset columns from CanCOLD, SPIROMICS etc and saves them in a separate sheet of a newly created excel file###

import openpyxl

def extract_and_save_column_names(datasets, output_file):
    # Create a new Excel workbook
    wb_output = openpyxl.Workbook()
    
    # Remove the default sheet created by Workbook()
    default_sheet = wb_output.active
    wb_output.remove(default_sheet)

    for dataset in datasets:
        input_file = dataset['input_file']
        sheet_name = dataset['sheet_name']
        output_sheet_name = dataset['output_sheet_name']
        
        # Load the input Excel workbook and select the specified sheet
        wb_input = openpyxl.load_workbook(input_file)
        sheet = wb_input[sheet_name]

        # Extract column names (assuming the first row contains the column names)
        column_names = [cell.value for cell in sheet[1]]

        # Add a new sheet to the output workbook with the specified sheet name
        output_sheet = wb_output.create_sheet(title=output_sheet_name)

        # Write the column names to the first column of the new sheet
        for row_num, column_name in enumerate(column_names, start=1):
            output_sheet.cell(row=row_num, column=1, value=column_name)
    
    # Save the new Excel workbook
    wb_output.save(output_file)

# Example usage
datasets = [
    {
        'input_file': 'Z:\\ra\\data_harmon\\cancold\\CanCOLD_Variable_List.xlsx',  # Path to the first input Excel file
        'sheet_name': 'Demo+WholeLungCT',       # Name of the sheet to extract column names from
        'output_sheet_name': 'v1_can_var'  # Name of the sheet in the output Excel file
    },
    {
        'input_file': 'Z:/ra/data_harmon/cancold/V2_CanCOLD_Variable_List.xlsx',  # Path to the second input Excel file
        'sheet_name': 'Sheet1',       # Name of the sheet to extract column names from
        'output_sheet_name': 'v2_can_var'  # Name of the sheet in the output Excel file
    },
    {
        'input_file': 'Z:/ra/data_harmon/cancold/V3_CanCOLD_Variable_List.xlsx',  # Path to the second input Excel file
        'sheet_name': 'Sheet1',       # Name of the sheet to extract column names from
        'output_sheet_name': 'v3_can_var'  # Name of the sheet in the output Excel file
    },
    {
        'input_file': 'Z:\\ra\\data_harmon\\spiromics\\SPIROMICS_All_Data.xlsx',  # Path to the second input Excel file
        'sheet_name': 'SPIROMICS_All_Data',       # Name of the sheet to extract column names from
        'output_sheet_name': 'v1_spiromics_var'  # Name of the sheet in the output Excel file
    }
]

output_file = 'var_col.xlsx'  # Path to save the output Excel file

extract_and_save_column_names(datasets, output_file)

